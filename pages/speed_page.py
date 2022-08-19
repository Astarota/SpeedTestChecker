from selenium.common.exceptions import ElementNotInteractableException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import NoSuchElementException
from .base_page import BasePage
from .locators import SpeedPageLocators
import time
import openpyxl as O

Excel_file="C:\\Users\\Dossan\\Desktop\\Results.xlsx"
Excel_worksheet="Лист1"

class SpeedPage(BasePage):
	def should_be_add_to_cart(self):
		self.should_be_go_click()
		self.should_be_download_number()
		self.should_be_upload_number()

	def should_be_go_click(self):
		btn = self.browser.find_element(*SpeedPageLocators.go_btn)
		btn.click()
		time.sleep(40)

	def should_be_close_window(self):
		try:
			btn=self.browser.find_element(*SpeedPageLocators.close_btn)
			btn.click()
		except(ElementNotInteractableException,StaleElementReferenceException,NoSuchElementException):
			return False
		return True

	def should_be_putted_in_excel_download_number(self,i):
		bool=self.is_element_present(*SpeedPageLocators.download_speed)
		if(bool==True):
			download_number = self.browser.find_element(*SpeedPageLocators.download_speed)
			print(download_number.text)
			wb=O.load_workbook(Excel_file)
			ws=wb[Excel_worksheet]
			ws.cell(i,2).value=download_number.text
			wb.save(Excel_file)
			wb.close()
		elif(bool==False):
			close_btn=self.browser.find_element(*SpeedPageLocators.close_fail_btn)
			close_btn.click()
			wb=O.load_workbook(Excel_file)
			ws=wb[Excel_worksheet]
			ws.cell(i,2).value="FAIL"
			wb.save(Excel_file)
			wb.close()

	def should_be_putted_in_excel_upload_number(self,i):
		bool=self.is_element_present(*SpeedPageLocators.upload_speed)
		if(bool==True):
			upload_number = self.browser.find_element(*SpeedPageLocators.upload_speed)
			print(upload_number.text)
			wb=O.load_workbook(Excel_file)
			ws=wb[Excel_worksheet]
			ws.cell(i,3).value=upload_number.text
			wb.save(Excel_file)
			wb.close()
		elif(bool==False):
			close_btn=self.browser.find_element(*SpeedPageLocators.close_fail_btn)
			close_btn.click()
			wb=O.load_workbook(Excel_file)
			ws=wb[Excel_worksheet]
			ws.cell(i,3).value="FAIL"
			wb.save(Excel_file)
			wb.close()

